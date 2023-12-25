from googletrans import Translator,LANGUAGES
from gtts import gTTS  
from openpyxl import Workbook, load_workbook
def chatgpt_copy_n_past(cardtype,phrase,word=''):
    if cardtype=="v":
      prompt=translate("Explain what this word")
      prompt+=f"_{word}_"
      prompt+=translate("means in this phrase and also in general")
      prompt+=f"_{phrase}_"
      return prompt 
    else:
      return translate(f"Translate this phrase to my mother tongue and explain it is meaning") +phrase         
def check_translation(cardtype,phrases,translated_phrases,
                      words='',translated_words='',
                      check_table_name='check_table.xlsx'):
    workbook_checking= Workbook()
    check_table=workbook_checking.active
    vocab_header=["ChatGPT"] 
    if cardtype=="v": vocab_header.append("Words","Translated Words")
    check_table.append(["Phrases","Translated_Phrases"]+ vocab_header)
    for i in range(len(phrases)):
      check_table["A"+str(i+2)]=phrases[i]
      check_table["B"+str(i+2)]=translated_phrases[i]
      check_table["C"+str(i+2)]=chatgpt_copy_n_past(cardtype,phrases[i],words[i])
      if cardtype=="v":
        check_table["D"+str(i+2)]=words[i]
        check_table["E"+str(i+2)]=translated_words[i]
    workbook_checking.save(check_table_name)
    workbook_checking.close()
def import_csv_table(file_path):
    from csv import DictReader
    with open(file_path,mode='r',encoding='utf16') as file:
        try:
            csv_dict=[x for x in DictReader(file)]
        except FileNotFoundError:
            return f"Error: File not found at {file_path}
        except Exception as e:
            return f"An error occurred: {e}"
    phrases=[x['Front'] for x in csv_dict if 'Front' in x]
    words=[x['Back'] for x in csv_dict if 'Back' in x]
    trans_phrases=[x['trans_phrases'] for x in csv_dict 
                        if 'trans_phrases' in x]
    trans_words=[x['trans_words'] for x in csv_dict 
                        if 'trans_words' in x]
    langs=[x['Lang'] for x in csv_dict if 'Lang' in x]
    return phrases,words,trans_phrases,trans_words,langs
def import_excel_table(file_path):
    from openpyxl import load_workbook
    try:
        workbook=load_workbook(file_path)
    except FileNotFoundError:
        return f"Error: File not found at {file_path}"
    except Exception as e:
        return f"An error occurred: {e}"
    sheet=workbook.active
    headers=[cell.value for cell in sheet[1]]
    dicts=[]
    for row in sheet.iter_rows(min_row=2,values_only=True):
        row_dict=dict(zip(headers,row))
        dicts.append(row_dict)
    workbook.close()
    phrases=[x['Front'] for x in dicts if 'Front' in x]
    words=[x['Back'] for x in dicts if 'Back' in x]
    trans_phrases=[x['trans_phrases'] for x in dicts 
                        if 'trans_phrases' in x]
    trans_words=[x['trans_words'] for x in dicts
                        if 'trans_words' in x]
    langs=[x['Lang'] for x in dicts if 'Lang' in x]
    return phrases,words,trans_phrases,trans_words,langs
def translate(message,translator=Translator(),source="en",chosen_language='pt'):
  return translator.translate(message, src=source,dest=chosen_language).text
def generate_audios(cardtype,langs,phrases,audio_path,deck_name,words=''):
  for j in range(len(phrases)):
        if cardtype=="v":
          audio_phrase = gTTS(f"{words[j]}.{phrases[j]}", lang=langs[j])
          audio_phrase.save(f"{audio_path}//{deck_name}phrase{j}.mp3")
          audio_word= gTTS(words[j], lang=langs[j])
          audio_word.save(f"{audio_path}//{deck_name}word{j}.mp3")
        elif cardtype in ["s","w"]:
          audio_phrase = gTTS(phrases[j], lang=langs[j])
          audio_phrase.save(f"{audio_path}//{deck_name}phrase{j}.mp3")
        elif cardtype=="p":
          audio = gTTS(phrases[j], lang=langs[j])
          audio.save(f"{audio_path}//{deck_name}pronunciation{j}.mp3")
def generate_translations(cardtype,langs,chosen_lang,phrases,words=''):
  translated_phrases=[translate(phrases[i],source=langs[i],chosen_language=chosen_lang) for i in range(len(phrases))]
  if cardtype=="v": translated_words=[translate(words[i],source=langs[i]) for i in range(len(phrases))]
  else: translated_words=['' for i in range(len(phrases))]
  return translated_phrases,translated_words