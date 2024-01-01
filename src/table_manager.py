def import_csv_table(file_path):
    '''Imports data from a CSV file containing a language learning table.

    Parameters:
    - file_path (str): The path to the CSV file.

    Returns:
    Tuple[List[str], List[str], List[str], List[str], List[str]]: A tuple containing lists of phrases, words,
    translated phrases, translated words, and languages extracted from the CSV file.

    Raises:
    - FileNotFoundError: If the specified file is not found.
    - Exception: If an error occurs during file reading or data extraction.

    Example:
    phrases, words, trans_phrases, trans_words, langs = import_csv_table('path/to/file.csv')
    '''
    from csv import DictReader
    with open(file_path,mode='r',encoding='utf16') as file:
        try:
            csv_dict=[x for x in DictReader(file)]
        except FileNotFoundError:
            return f"Error: File not found at {file_path}
        except Exception as e:
            return f"An error occurred: {e}"
    phrases=[x['Front'] for x in csv_dict if 'Front' in x]
    words=[x['Back'] for x in csv_dict if 'Back' in x]
    trans_phrases=[x['trans_phrases'] for x in csv_dict 
                        if 'trans_phrases' in x]
    trans_words=[x['trans_words'] for x in csv_dict 
                        if 'trans_words' in x]
    langs=[x['Lang'] for x in csv_dict if 'Lang' in x]
    return phrases,words,trans_phrases,trans_words,langs
def import_excel_table(file_path):
    '''Opens an Excel file and extracts data from the active sheet.

    Parameters:
    - file_path (str): The path to the Excel file.

    Returns:
    Tuple[List[str], List[str], List[str], List[str], List[str]]: A tuple containing lists of phrases, words,
    translated phrases, translated words, and languages extracted from the Excel file.

    Raises:
    - FileNotFoundError: If the specified file is not found.
    - Exception: If an error occurs during file reading or data extraction.
    
    Example:
    phrases, words, trans_phrases, trans_words, langs = import_excel_table('path/to/file.xlsx')
    '''
    from openpyxl import load_workbook
    try:
        workbook=load_workbook(file_path)
    except FileNotFoundError:
        return f"Error: File not found at {file_path}"
    except Exception as e:
        return f"An error occurred: {e}"
    sheet=workbook.active
    headers=[cell.value for cell in sheet[1]]
    dicts=[]
    for row in sheet.iter_rows(min_row=2,values_only=True):
        row_dict=dict(zip(headers,row))
        dicts.append(row_dict)
    workbook.close()
    phrases=[x['Front'] for x in dicts if 'Front' in x]
    words=[x['Back'] for x in dicts if 'Back' in x]
    trans_phrases=[x['trans_phrases'] for x in dicts 
                        if 'trans_phrases' in x]
    trans_words=[x['trans_words'] for x in dicts
                        if 'trans_words' in x]
    langs=[x['Lang'] for x in dicts if 'Lang' in x]
    return phrases,words,trans_phrases,trans_words,langs